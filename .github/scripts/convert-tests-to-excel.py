#!/usr/bin/env python3
"""
Script để convert kết quả test từ TRX format sang Excel format
theo template của BlindTreasure project
"""

import argparse
import xml.etree.ElementTree as ET
import pandas as pd
from datetime import datetime, timezone
import os
import glob
import json
import sys
from pathlib import Path
import re

class TestResultConverter:
    def __init__(self, test_results_dir, reports_dir, output_file, project_name, version, branch, build_number):
        self.test_results_dir = Path(test_results_dir)
        self.reports_dir = Path(reports_dir) if reports_dir else None
        self.output_file = output_file
        self.project_name = project_name
        self.version = version
        self.branch = branch
        self.build_number = build_number
        self.current_time = datetime.now(timezone.utc)
        
    def parse_trx_files(self):
        """Parse tất cả file TRX và extract test results"""
        trx_files = list(self.test_results_dir.glob("**/*.trx"))
        
        if not trx_files:
            print(f"❌ Không tìm thấy file TRX nào trong {self.test_results_dir}")
            return []
        
        all_test_results = []
        
        for trx_file in trx_files:
            print(f"📄 Đang xử lý file: {trx_file}")
            try:
                test_results = self._parse_single_trx(trx_file)
                all_test_results.extend(test_results)
                print(f"✅ Đã parse {len(test_results)} test cases từ {trx_file.name}")
            except Exception as e:
                print(f"❌ Lỗi khi parse file {trx_file}: {e}")
                continue
        
        return all_test_results
    
    def _parse_single_trx(self, trx_file_path):
        """Parse một file TRX duy nhất"""
        tree = ET.parse(trx_file_path)
        root = tree.getroot()
        
        # Namespace của TRX file (Microsoft Test Results)
        namespaces = {
            'ns': 'http://microsoft.com/schemas/VisualStudio/TeamTest/2010'
        }
        
        test_results = []
        test_id_counter = 1
        
        # Parse UnitTestResult elements
        for result in root.findall('.//ns:UnitTestResult', namespaces):
            test_name = result.get('testName', '')
            outcome = result.get('outcome', 'Unknown')
            duration = result.get('duration', '00:00:00')
            start_time = result.get('startTime', '')
            end_time = result.get('endTime', '')
            
            # Tạo Test Case ID theo format TC{number:03d}
            test_case_id = f"TC{test_id_counter:03d}"
            
            # Extract class name và method name từ test name
            class_name, method_name = self._extract_class_and_method(test_name)
            
            # Parse error information nếu có
            error_message = ""
            stack_trace = ""
            
            error_info = result.find('.//ns:ErrorInfo', namespaces)
            if error_info is not None:
                message_elem = error_info.find('.//ns:Message', namespaces)
                if message_elem is not None and message_elem.text:
                    error_message = message_elem.text.strip()
                
                stack_trace_elem = error_info.find('.//ns:StackTrace', namespaces)
                if stack_trace_elem is not None and stack_trace_elem.text:
                    stack_trace = stack_trace_elem.text.strip()
            
            # Generate test description và procedure dựa trên method name
            description, procedure, expected_result = self._generate_test_details(method_name, class_name)
            
            # Determine test status
            test_status = self._map_outcome_to_status(outcome)
            
            test_data = {
                'Test Case ID': test_case_id,
                'Test Case Description': description,
                'Test Case Procedure': procedure,
                'Expected Results': expected_result,
                'Pre-conditions': 'System setup and dependencies initialized',
                'Round 1': test_status,
                'Test date': self._format_date(start_time) if start_time else self.current_time.strftime('%d/%m/%Y'),
                'Tester': 'GitHub Actions CI/CD',
                'Round 2': '',
                'Test date.1': '',  # Second Test date column
                'Tester.1': '',     # Second Tester column
                'Round 3': '',
                'Test date.2': '',  # Third Test date column
                'Tester.2': '',     # Third Tester column
                'Note': self._generate_note(outcome, error_message, duration),
                # Additional technical details for reference
                '_technical_details': {
                    'full_test_name': test_name,
                    'class_name': class_name,
                    'method_name': method_name,
                    'outcome': outcome,
                    'duration': duration,
                    'start_time': start_time,
                    'end_time': end_time,
                    'error_message': error_message,
                    'stack_trace': stack_trace
                }
            }
            
            test_results.append(test_data)
            test_id_counter += 1
        
        return test_results
    
    def _extract_class_and_method(self, test_name):
        """Extract class name và method name từ full test name"""
        # Pattern: Namespace.ClassName.MethodName hoặc ClassName.MethodName
        parts = test_name.split('.')
        if len(parts) >= 2:
            method_name = parts[-1]
            class_name = parts[-2]
        else:
            method_name = test_name
            class_name = "Unknown"
        
        return class_name, method_name
    
    def _generate_test_details(self, method_name, class_name):
        """Generate test description, procedure, và expected result dựa trên method name"""
        
        # Convert CamelCase to readable text
        readable_name = re.sub(r'([A-Z])', r' \1', method_name).strip()
        readable_name = re.sub(r'^Test\s*', '', readable_name, flags=re.IGNORECASE)
        
        # Common test patterns
        if 'should' in method_name.lower():
            description = f"Verify that {readable_name.lower()}"
        elif 'when' in method_name.lower():
            description = f"Test behavior {readable_name.lower()}"
        elif 'get' in method_name.lower():
            description = f"Verify {readable_name.lower()} functionality"
        elif 'create' in method_name.lower() or 'add' in method_name.lower():
            description = f"Test {readable_name.lower()} operation"
        elif 'update' in method_name.lower() or 'edit' in method_name.lower():
            description = f"Verify {readable_name.lower()} functionality"
        elif 'delete' in method_name.lower() or 'remove' in method_name.lower():
            description = f"Test {readable_name.lower()} operation"
        else:
            description = f"Verify {readable_name.lower()}"
        
        # Generate procedure
        procedure = f"""1. Setup test environment and dependencies
2. Initialize {class_name} service/controller
3. Execute {method_name} method
4. Verify the results and assertions"""
        
        # Generate expected result
        if 'exception' in method_name.lower() or 'error' in method_name.lower():
            expected_result = "Expected exception or error should be thrown with appropriate message"
        elif 'return' in method_name.lower() or 'get' in method_name.lower():
            expected_result = "Method should return expected value/object with correct data"
        elif 'success' in method_name.lower():
            expected_result = "Operation should complete successfully without errors"
        elif 'fail' in method_name.lower() or 'invalid' in method_name.lower():
            expected_result = "Operation should fail gracefully with appropriate error handling"
        else:
            expected_result = "All assertions should pass and method should execute as expected"
        
        return description, procedure, expected_result
    
    def _map_outcome_to_status(self, outcome):
        """Map TRX outcome to test status"""
        outcome_mapping = {
            'Passed': 'Passed',
            'Failed': 'Failed',
            'NotExecuted': 'Skipped',
            'Inconclusive': 'Inconclusive',
            'Error': 'Error',
            'Timeout': 'Timeout',
            'Aborted': 'Aborted'
        }
        
        return outcome_mapping.get(outcome, 'Unknown')
    
    def _format_date(self, date_string):
        """Format date string to DD/MM/YYYY"""
        try:
            if date_string:
                # Parse ISO format: 2024-07-26T10:30:45.123Z
                dt = datetime.fromisoformat(date_string.replace('Z', '+00:00'))
                return dt.strftime('%d/%m/%Y')
        except:
            pass
        return self.current_time.strftime('%d/%m/%Y')
    
    def _generate_note(self, outcome, error_message, duration):
        """Generate note based on test outcome"""
        notes = []
        
        # Add duration info
        if duration and duration != '00:00:00':
            notes.append(f"Duration: {duration}")
        
        # Add outcome specific notes
        if outcome == 'Failed' and error_message:
            # Truncate long error messages
            short_error = error_message[:100] + "..." if len(error_message) > 100 else error_message
            notes.append(f"Error: {short_error}")
        elif outcome == 'Passed':
            notes.append("All assertions passed")
        elif outcome == 'NotExecuted':
            notes.append("Test was skipped")
        
        return " | ".join(notes) if notes else ""
    
    def generate_summary_statistics(self, test_results):
        """Tạo thống kê tổng quan"""
        total_tests = len(test_results)
        passed = sum(1 for t in test_results if t['Round 1'] == 'Passed')
        failed = sum(1 for t in test_results if t['Round 1'] == 'Failed')
        skipped = sum(1 for t in test_results if t['Round 1'] == 'Skipped')
        other = total_tests - passed - failed - skipped
        
        pass_rate = (passed / total_tests * 100) if total_tests > 0 else 0
        
        return {
            'Project Name': self.project_name,
            'Version': self.version,
            'Branch': self.branch,
            'Build Number': self.build_number,
            'Test Execution Date': self.current_time.strftime('%d/%m/%Y %H:%M:%S UTC'),
            'Total Test Cases': total_tests,
            'Passed': passed,
            'Failed': failed,
            'Skipped': skipped,
            'Other': other,
            'Pass Rate (%)': f"{pass_rate:.2f}%",
            'Environment': 'GitHub Actions CI/CD'
        }
    
    def load_coverage_data(self):
        """Load code coverage data nếu có"""
        coverage_data = None
        
        if self.reports_dir:
            # Tìm file coverage JSON
            coverage_files = list(self.reports_dir.glob("**/Summary.json"))
            if coverage_files:
                try:
                    with open(coverage_files[0], 'r') as f:
                        coverage_data = json.load(f)
                    print(f"✅ Đã load coverage data từ {coverage_files[0]}")
                except Exception as e:
                    print(f"⚠️ Không thể load coverage data: {e}")
        
        return coverage_data
    
    def create_coverage_summary(self, coverage_data):
        """Tạo summary cho code coverage"""
        if not coverage_data:
            return {
                'Line Coverage': 'N/A',
                'Branch Coverage': 'N/A',
                'Method Coverage': 'N/A',
                'Class Coverage': 'N/A'
            }
        
        try:
            summary = coverage_data.get('summary', {})
            return {
                'Line Coverage': f"{summary.get('linecoverage', 0) * 100:.2f}%",
                'Branch Coverage': f"{summary.get('branchcoverage', 0) * 100:.2f}%",
                'Method Coverage': f"{summary.get('methodcoverage', 0) * 100:.2f}%",
                'Class Coverage': f"{summary.get('classcoverage', 0) * 100:.2f}%"
            }
        except:
            return {
                'Line Coverage': 'N/A',
                'Branch Coverage': 'N/A', 
                'Method Coverage': 'N/A',
                'Class Coverage': 'N/A'
            }
    
    def export_to_excel(self):
        """Export test results to Excel với multiple sheets"""
        print(f"🚀 Bắt đầu generate Excel report...")
        
        # Parse test results
        test_results = self.parse_trx_files()
        
        if not test_results:
            print("❌ Không có kết quả test nào để export!")
            return False
        
        # Load coverage data
        coverage_data = self.load_coverage_data()
        
        # Tạo DataFrames
        df_tests = pd.DataFrame(test_results)
        
        # Remove technical details column for Excel export
        if '_technical_details' in df_tests.columns:
            df_tests = df_tests.drop('_technical_details', axis=1)
        
        # Tạo summary statistics
        summary_stats = self.generate_summary_statistics(test_results)
        df_summary = pd.DataFrame([summary_stats])
        
        # Tạo coverage summary
        coverage_summary = self.create_coverage_summary(coverage_data)
        df_coverage = pd.DataFrame([coverage_summary])
        
        # Filter failed tests
        df_failed = df_tests[df_tests['Round 1'] == 'Failed'].copy()
        
        print(f"📊 Exporting {len(test_results)} test results to Excel...")
        
        # Export to Excel với multiple sheets
        try:
            with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
                # Sheet 1: Test Cases (Main sheet theo template)
                df_tests.to_excel(writer, sheet_name='Test Cases', index=False, startrow=1)
                
                # Sheet 2: Summary
                df_summary.to_excel(writer, sheet_name='Summary', index=False)
                
                # Sheet 3: Code Coverage
                df_coverage.to_excel(writer, sheet_name='Code Coverage', index=False)
                
                # Sheet 4: Failed Tests (nếu có)
                if not df_failed.empty:
                    df_failed.to_excel(writer, sheet_name='Failed Tests', index=False)
                
                # Format các sheets
                self._format_excel_sheets(writer, df_tests, df_summary, df_coverage, df_failed)
            
            print(f"✅ Đã export thành công ra file: {self.output_file}")
            self._print_summary(summary_stats)
            
            return True
            
        except Exception as e:
            print(f"❌ Lỗi khi export Excel: {e}")
            return False
    
    def _format_excel_sheets(self, writer, df_tests, df_summary, df_coverage, df_failed):
        """Format Excel sheets cho đẹp"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Colors
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        passed_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
failed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
       skipped_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
       
       header_font = Font(color="FFFFFF", bold=True)
       border = Border(
           left=Side(style='thin'),
           right=Side(style='thin'),
           top=Side(style='thin'),
           bottom=Side(style='thin')
       )
       
       # Format Test Cases sheet
       test_sheet = writer.sheets['Test Cases']
       
       # Add title
       test_sheet['A1'] = f"{self.project_name} - Unit Test Report"
       test_sheet['A1'].font = Font(size=16, bold=True)
       test_sheet.merge_cells('A1:N1')
       
       # Format headers (row 2)
       for col_num, column_title in enumerate(df_tests.columns, 1):
           cell = test_sheet.cell(row=2, column=col_num)
           cell.fill = header_fill
           cell.font = header_font
           cell.alignment = Alignment(horizontal='center', vertical='center')
           cell.border = border
       
       # Auto-adjust column widths và format data
       for col_num, column in enumerate(df_tests.columns, 1):
           column_letter = get_column_letter(col_num)
           
           # Set column width based on content
           if column in ['Test Case ID']:
               test_sheet.column_dimensions[column_letter].width = 12
           elif column in ['Test Case Description']:
               test_sheet.column_dimensions[column_letter].width = 40
           elif column in ['Test Case Procedure', 'Expected Results']:
               test_sheet.column_dimensions[column_letter].width = 50
           elif column in ['Round 1', 'Round 2', 'Round 3']:
               test_sheet.column_dimensions[column_letter].width = 10
           elif column in ['Test date', 'Test date.1', 'Test date.2']:
               test_sheet.column_dimensions[column_letter].width = 12
           elif column in ['Tester', 'Tester.1', 'Tester.2']:
               test_sheet.column_dimensions[column_letter].width = 15
           elif column == 'Note':
               test_sheet.column_dimensions[column_letter].width = 30
           else:
               test_sheet.column_dimensions[column_letter].width = 15
       
       # Format data rows with conditional formatting
       for row_num in range(3, len(df_tests) + 3):
           round1_col = None
           for col_num, column in enumerate(df_tests.columns, 1):
               cell = test_sheet.cell(row=row_num, column=col_num)
               cell.border = border
               cell.alignment = Alignment(vertical='top', wrap_text=True)
               
               if column == 'Round 1':
                   round1_col = col_num
                   # Color based on test result
                   if cell.value == 'Passed':
                       cell.fill = passed_fill
                   elif cell.value == 'Failed':
                       cell.fill = failed_fill
                   elif cell.value == 'Skipped':
                       cell.fill = skipped_fill
       
       # Format Summary sheet
       summary_sheet = writer.sheets['Summary']
       summary_sheet['A1'] = f"{self.project_name} - Test Execution Summary"
       summary_sheet['A1'].font = Font(size=14, bold=True)
       summary_sheet.merge_cells('A1:B1')
       
       # Format summary headers
       for row in range(2, len(df_summary) + 2):
           for col in range(1, len(df_summary.columns) + 1):
               cell = summary_sheet.cell(row=row, column=col)
               if col == 1:  # Header column
                   cell.fill = header_fill
                   cell.font = header_font
               cell.border = border
               cell.alignment = Alignment(vertical='center')
       
       # Auto-adjust summary columns
       for col_num in range(1, len(df_summary.columns) + 1):
           column_letter = get_column_letter(col_num)
           summary_sheet.column_dimensions[column_letter].width = 25
       
       # Format Coverage sheet
       coverage_sheet = writer.sheets['Code Coverage']
       coverage_sheet['A1'] = f"{self.project_name} - Code Coverage Report"
       coverage_sheet['A1'].font = Font(size=14, bold=True)
       coverage_sheet.merge_cells('A1:B1')
       
       # Format coverage headers
       for row in range(2, len(df_coverage) + 2):
           for col in range(1, len(df_coverage.columns) + 1):
               cell = coverage_sheet.cell(row=row, column=col)
               if col == 1:  # Header column
                   cell.fill = header_fill
                   cell.font = header_font
               cell.border = border
               cell.alignment = Alignment(vertical='center')
       
       # Auto-adjust coverage columns
       for col_num in range(1, len(df_coverage.columns) + 1):
           column_letter = get_column_letter(col_num)
           coverage_sheet.column_dimensions[column_letter].width = 20
       
       # Format Failed Tests sheet if exists
       if not df_failed.empty:
           failed_sheet = writer.sheets['Failed Tests']
           failed_sheet['A1'] = f"{self.project_name} - Failed Tests Details"
           failed_sheet['A1'].font = Font(size=14, bold=True, color="FF0000")
           failed_sheet.merge_cells('A1:N1')
           
           # Format failed tests similar to main sheet
           for col_num, column_title in enumerate(df_failed.columns, 1):
               cell = failed_sheet.cell(row=2, column=col_num)
               cell.fill = header_fill
               cell.font = header_font
               cell.alignment = Alignment(horizontal='center', vertical='center')
               cell.border = border
           
           # Format failed test data rows
           for row_num in range(3, len(df_failed) + 3):
               for col_num, column in enumerate(df_failed.columns, 1):
                   cell = failed_sheet.cell(row=row_num, column=col_num)
                   cell.border = border
                   cell.alignment = Alignment(vertical='top', wrap_text=True)
                   if column == 'Round 1':
                       cell.fill = failed_fill
   
   def _print_summary(self, summary_stats):
       """In ra console summary statistics"""
       print("\n" + "="*50)
       print(f"📊 TEST EXECUTION SUMMARY")
       print("="*50)
       print(f"Project: {summary_stats['Project Name']}")
       print(f"Version: {summary_stats['Version']}")
       print(f"Branch: {summary_stats['Branch']}")
       print(f"Build: {summary_stats['Build Number']}")
       print(f"Date: {summary_stats['Test Execution Date']}")
       print("-"*50)
       print(f"Total Tests: {summary_stats['Total Test Cases']}")
       print(f"✅ Passed: {summary_stats['Passed']}")
       print(f"❌ Failed: {summary_stats['Failed']}")
       print(f"⏭️ Skipped: {summary_stats['Skipped']}")
       print(f"📊 Pass Rate: {summary_stats['Pass Rate (%)']}")
       print("="*50)

def main():
   parser = argparse.ArgumentParser(description='Convert .NET Unit Test TRX results to Excel format')
   parser.add_argument('--test-results-dir', required=True, help='Directory containing TRX files')
   parser.add_argument('--reports-dir', help='Directory containing coverage reports')
   parser.add_argument('--output-file', required=True, help='Output Excel file name')
   parser.add_argument('--project-name', default='BlindTreasure API', help='Project name')
   parser.add_argument('--version', required=True, help='Version/commit hash')
   parser.add_argument('--branch', required=True, help='Git branch name')
   parser.add_argument('--build-number', required=True, help='Build number')
   
   args = parser.parse_args()
   
   print(f"🚀 Starting Excel report generation...")
   print(f"📁 Test Results Dir: {args.test_results_dir}")
   print(f"📁 Reports Dir: {args.reports_dir}")
   print(f"📄 Output File: {args.output_file}")
   print(f"🏷️ Project: {args.project_name}")
   print(f"🔖 Version: {args.version}")
   print(f"🌿 Branch: {args.branch}")
   print(f"🔢 Build: {args.build_number}")
   
   converter = TestResultConverter(
       test_results_dir=args.test_results_dir,
       reports_dir=args.reports_dir,
       output_file=args.output_file,
       project_name=args.project_name,
       version=args.version,
       branch=args.branch,
       build_number=args.build_number
   )
   
   success = converter.export_to_excel()
   
   if success:
       print("🎉 Excel report generation completed successfully!")
       sys.exit(0)
   else:
       print("💥 Excel report generation failed!")
       sys.exit(1)

if __name__ == "__main__":
   main()